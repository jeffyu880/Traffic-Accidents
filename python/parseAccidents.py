import openpyxl as xl
import os

print(os.path.abspath(os.curdir))
wb = xl.load_workbook('US_Accidents_Dec20_Final_test.xlsx')
#wb = xl.load_workbook('TestLocations.xlsx')
accidentDataSheet = wb['AccidentData']
recordNumberSheet = wb['Sheet1']

citiestoDelete = []

# this loops through the sorted cities to see if there are less than 100 accidents in the city
# and appends it to a list of cities to delete
for row in range(1, recordNumberSheet.max_row+1):
    if recordNumberSheet.cell(row, 4).value <= 100:
        citiestoDelete.append(recordNumberSheet.cell(row, 3).value)
        #print("adding: city " + str(recordNumberSheet.cell(row, 3).value) + " accidents " + str(recordNumberSheet.cell(row, 4).value))

    
for x in range(len(citiestoDelete)):
    print(citiestoDelete[x])
    for row in range (1, 100):
        #print(accidentDataSheet.cell(row, 14).value)
        if citiestoDelete[x] == accidentDataSheet.cell(row, 14).value:
            print("deleting " + accidentDataSheet.cell(row, 14).value)
            #accidentDataSheet.delete_rows(row)

wb.save('US_Accidents_Dec20_Final_Parsed.xlsx')
#wb.save('TestLocations1.xlsx')
